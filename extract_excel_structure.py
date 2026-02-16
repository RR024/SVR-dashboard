from openpyxl import load_workbook
import sys

file_path = "srcs/6.Matrl Calcualtion 25-26.xlsx"

print("\n" + "="*70)
print("MATERIAL CALCULATION EXCEL - SHEET & COLUMN EXTRACTION")
print("="*70 + "\n")

try:
    wb = load_workbook(file_path, data_only=True, read_only=True)
    
    print(f"📁 File: {file_path}")
    print(f"📊 Total Sheets: {len(wb.sheetnames)}\n")
    
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"\n{'─'*70}")
        print(f"Sheet {idx}: {sheet_name}")
        print(f"{'─'*70}")
        
        ws = wb[sheet_name]
        
        # Get dimensions
        print(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Get column headers (first row)
        print(f"\nColumns:")
        headers = []
        for col in range(1, min(ws.max_column + 1, 30)):  # Max 30 columns to display
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value))
                print(f"  {col}. {cell_value}")
        
        # Show first 3 data rows
        print(f"\nFirst 3 Data Rows:")
        for row in range(2, min(5, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(len(headers) + 1, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else "")
            print(f"  Row {row}: {row_data[:5]}")  # Show first 5 values
    
    wb.close()
    print(f"\n{'='*70}")
    print("EXTRACTION COMPLETE")
    print(f"{'='*70}\n")
    
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()
